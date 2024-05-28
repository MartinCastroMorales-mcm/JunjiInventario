from flask import request, flash, render_template, url_for, redirect, Blueprint, session
from db import mysql
from funciones import getPerPage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from cuentas import loguear_requerido, administrador_requerido

equipo = Blueprint("equipo", __name__, template_folder="app/templates")


# envia datos al formulario y tabla de equipo CAMBIA FK_IDCODIGO_PROVEEDOR
@equipo.route("/equipo")
@equipo.route("/equipo/<page>")
@loguear_requerido
def Equipo(page=1):
    page = int(page)
    perpage = getPerPage()
    offset = (int(page) - 1) * perpage
    #solo funciona con connect no con connect
    #si funciona con connection. parece que era algo de la maquina virtual
    #elimine la maquina virtual y ahora funciona
    cur = mysql.connection.cursor() #ahora connect funciona pero no connection ¿?
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    cur.execute(""" 
    SELECT *
    FROM super_equipo
    LIMIT {} OFFSET {}

    """.format(
            perpage, offset
        )
    )
    equipos = cur.fetchall()
    modelos_por_tipo = cur.fetchall()
    cur.execute("SELECT * FROM tipo_equipo")
    tipo_equipo = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    _data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("""
    SELECT *
    FROM marca_equipo
                """)
    marcas = cur.fetchall()
    print(marcas)

    modelos_por_tipo = {

    }
    for tipo in tipo_equipo:
        query = """
        SELECT *
        FROM modelo_equipo me
        WHERE me.idTipo_Equipo = {}
""".format(str(tipo['idTipo_equipo']))
        #print(query)
        cur.execute("""
        SELECT *
        FROM modelo_equipo me
        WHERE me.idTipo_Equipo = %s
            """, (tipo['idTipo_equipo'],))
        modelo_tipo = cur.fetchall()
        modelos_por_tipo[tipo['idTipo_equipo']] = modelo_tipo


    #print("tipos de equipo ############")
    #print(tipoe_data)
    return render_template(
        "equipo.html",
        equipo=equipos,
        tipo_equipo=tipo_equipo,
        marcas_equipo=marcas,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modelos_por_tipo,
        page=page,
        lastpage=page < (total / perpage) + 1,
        session=session
    )


# agrega registro para id
@equipo.route("/add_equipo", methods=["POST"])
@administrador_requerido
def add_equipo():
    if request.method == "POST":
        codigo_inventario = request.form["codigo_inventario"]
        numero_serie = request.form["numero_serie"]
        observacion_equipo = request.form["observacion_equipo"]
        codigoproveedor = request.form["codigoproveedor"]
        mac = request.form["mac"]
        imei = request.form["imei"]
        numero = request.form["numero"]
        #nombre_tipo_equipo = request.form["nombre_tipo_equipo"]
        #nombre_estado_equipo = request.form["nombre_estado_equipo"]
        codigo_Unidad = request.form["codigo_Unidad"]
        nombre_orden_compra = request.form["nombre_orden_compra"]
        nombre_modelo_equipo = request.form["nombre_modelo_equipo"]
        try:
            cur = mysql.connection.cursor()
            cur.execute(
                """ INSERT INTO equipo (
                    Cod_inventarioEquipo, 
                    Num_serieEquipo, 
                    ObservacionEquipo,
                    codigoproveedor_equipo,
                    macEquipo,
                    imeiEquipo, 
                    numerotelefonicoEquipo, 
                    idEstado_Equipo, 
                    idUnidad, 
                    idOrden_compra, 
                    idModelo_equipo) 
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s)
            """,
                (
                    codigo_inventario,
                    numero_serie,
                    observacion_equipo,
                    codigoproveedor,
                    mac,
                    imei,
                    numero,
                    3,
                    codigo_Unidad,
                    nombre_orden_compra,
                    nombre_modelo_equipo,
                ),
            )
            mysql.connection.commit()
            flash("Equipo agregado correctamente")
            return redirect(url_for("equipo.Equipo"))
        except Exception as e:
            print("exception agregar equipo")
            flash(e.args[1])
            return redirect(url_for("equipo.Equipo"))


# envia datos al formulario editar segun id
@equipo.route("/edit_equipo/<id>", methods=["POST", "GET"])
@administrador_requerido
def edit_equipo(id):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    try:
        cur = mysql.connection.cursor()
        cur.execute(
            """ 
           SELECT *
        FROM equipo e
        INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
        INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
        INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
        INNER JOIN unidad u on u.idUnidad = e.idUnidad
        INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
        WHERE idEquipo = %s
        """,
            (id,),
        )
        data = cur.fetchall()
        cur.execute("SELECT * FROM tipo_equipo")
        tipoe_data = cur.fetchall()
        cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
        estadoe_data = cur.fetchall()
        cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
        ubi_data = cur.fetchall()
        cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
        ordenc_data = cur.fetchall()
        cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
        modeloe_data = cur.fetchall()
        return render_template(
            "editEquipo.html",
            equipo=data[0],
            tipo_equipo=tipoe_data,
            estado_equipo=estadoe_data,
            orden_compra=ordenc_data,
            Unidad=ubi_data,
            modelo_equipo=modeloe_data,
        )
    except Exception as e:
        flash(e.args[1])
        return redirect(url_for("equipo.Equipo"))


# actualiza registro a traves de id correspondiente
@equipo.route("/update_equipo/<id>", methods=["POST"])
@administrador_requerido
def update_equipo(id):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    if request.method == "POST":
        codigo_inventario = request.form["codigo_inventario"]
        numero_serie = request.form["numero_serie"]
        observacion_equipo = request.form["observacion_equipo"]
        codigoproveedor = request.form["codigoproveedor"]
        mac = request.form["mac"]
        imei = request.form["imei"]
        numero = request.form["numero"]
        nombre_estado_equipo = request.form["nombre_estado_equipo"]
        codigo_Unidad = request.form["codigo_Unidad"]
        nombre_orden_compra = request.form["nombre_orden_compra"]
        nombre_modelo_equipo = request.form["nombre_modelo_equipo"]
        try:
            cur = mysql.connection.cursor()
            cur.execute(
                """
            UPDATE equipo
            SET Cod_inventarioEquipo = %s,
                Num_serieEquipo = %s,
                ObservacionEquipo = %s,
                codigoproveedor_equipo =%s,
                macEquipo=%s,
                imeiEquipo=%s,
                numerotelefonicoEquipo=%s,
                idEstado_Equipo = %s,
                idUnidad = %s,
                idOrden_compra = %s,
                idModelo_equipo = %s
                WHERE idEquipo = %s
            """,
                (
                    codigo_inventario,
                    numero_serie,
                    observacion_equipo,
                    codigoproveedor,
                    mac,
                    imei,
                    numero,
                    nombre_estado_equipo,
                    codigo_Unidad,
                    nombre_orden_compra,
                    nombre_modelo_equipo,
                    id,
                ),
            )
            mysql.connection.commit()
            flash("Equipo actualizado correctamente")
            return redirect(url_for("equipo.Equipo"))
        except Exception as e:
            flash(e.args[1])
            return redirect(url_for("equipo.Equipo"))


# elimina registro a traves de id correspondiente
@equipo.route("/delete_equipo/<id>", methods=["POST", "GET"])
@administrador_requerido
def delete_equipo(id):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    try:
        cur = mysql.connection.cursor()
        cur.execute("DELETE FROM equipo WHERE idEquipo = %s", (id,))
        mysql.connection.commit()
        flash("Equipo eliminado correctamente")
        return redirect(url_for("equipo.Equipo"))
    except Exception as e:
        flash(e.args[1])
        return redirect(url_for("equipo.Equipo"))


@equipo.route("/mostrar_asociados_traslado/<idTraslado>")
@loguear_requerido
def mostrar_asociados_traslado(idTraslado):
    page = 1
    perpage = 200
    offset = (page - 1) * perpage
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    cur.execute(
        """ 
    SELECT e.idEquipo, e.Cod_inventarioEquipo, e.Num_serieEquipo, e.ObservacionEquipo, 
    e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, e.numerotelefonicoEquipo,
    e.idEstado_Equipo, e.idUnidad, e.idOrden_compra, e.idModelo_equipo,te.idTipo_equipo, 
    te.nombreTipo_Equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
    moe.idModelo_equipo, moe.nombreModeloequipo
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    INNER JOIN traslacion tras on tras.idEquipo = e.idEquipo
    WHERE tras.idTraslado = %s
    LIMIT %s OFFSET %s
    """,
        (idTraslado, perpage, offset),
    )
    data = cur.fetchall()
    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    return render_template(
        "equipo.html",
        equipo=data,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=page,
        lastpage=False,
    )


@equipo.route("/mostrar_asociados_unidad/<idUnidad>")
@equipo.route("/mostrar_asociados_unidad/<idUnidad>/<page>")
@loguear_requerido
def mostrar_asociados_unidad(idUnidad, page=1):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    page = int(page)
    page = 1
    perpage = 200  # getPerPage()
    offset = (page - 1) * perpage
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    cur.execute(
        """ 
    SELECT e.idEquipo, e.Cod_inventarioEquipo, e.Num_serieEquipo, e.ObservacionEquipo, 
    e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, e.numerotelefonicoEquipo,
    e.idEstado_Equipo, e.idUnidad, e.idOrden_compra, e.idModelo_equipo,te.idTipo_equipo, 
    te.nombreTipo_Equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, 
    u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
    moe.idModelo_equipo, moe.nombreModeloequipo
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    WHERE e.idUnidad = %s
    LIMIT %s OFFSET %s
    """,
        (
            idUnidad,
            perpage,
            offset,
        ),
    )
    data = cur.fetchall()
    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    return render_template(
        "equipo.html",
        equipo=data,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=page,
        lastpage=False,
    )

@equipo.route("/mostrar_asociados_funcionario/<rutFuncionario>")
@equipo.route("/mostrar_asociados_funcionario/<rutFuncionario>/<page>")
@loguear_requerido
def mostrar_asociados_funcionario(rutFuncionario, page=1):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    page = int(page)
    page = 1
    perpage = 200 #porque ningun funcionario tendra tantos 
                    #equipos que la paginacion sea nesesaria
    offset = (page - 1) * perpage
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    #encontar la fecha de la ultima asignacion
    cur.execute("""
                SELECT *
                FROM asignacion a
                WHERE a.rutFuncionario = %s
                ORDER BY a.fecha_inicioAsignacion DESC
                """, (rutFuncionario,))
    asignaciones = cur.fetchall()
    print("########################")
    print(asignaciones)
    if(len(asignaciones) == 0):
        
        return render_template(
            "equipo.html",
            equipo=(),
            tipo_equipo=tipoe_data,
            estado_equipo=estadoe_data,
            orden_compra=ordenc_data,
            Unidad=ubi_data,
            modelo_equipo=modeloe_data,
            page=page,
            lastpage=False,
        )
    ultimaAsignacion = asignaciones[0]
    primeraAsignacion = asignaciones[-1] #-1 deberia dar el ultimo elemento

    cur.execute(""" 
    SELECT *
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    INNER JOIN equipo_asignacion ea on ea.idEquipo = e.idEquipo
    INNER JOIN asignacion a on a.idAsignacion = ea.idAsignacion
    INNER JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    WHERE a.idAsignacion = %s AND f.rutFuncionario = %s
    LIMIT %s OFFSET %s
    """,
        (
            ultimaAsignacion['idAsignacion'],
            rutFuncionario,
            perpage,
            offset,
        ),
    )
    data = cur.fetchall()
    return render_template(
        "equipo.html",
        equipo=data,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=page,
        lastpage=False,
    )


@equipo.route("/equipo_detalles/<idEquipo>")
@loguear_requerido
def equipo_detalles(idEquipo):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    cur = mysql.connection.cursor()
    #Como funcionaria con la asignacion cambiada ¿?
    #Cuando se añadan las asignaciones y devoluciones agregar funcionario como nombre
    #TODO: Revisar que hacer con las observaciones de Traslado, 
    #Revisar que hacer con observacion de Devolucion
    cur.execute("""
                SELECT i.fechaIncidencia as fecha, i.idIncidencia as id,
                    "Incidencia" as evento, i.observacionIncidencia as observacion,
                    i.nombreIncidencia as nombre
                FROM incidencia i
                WHERE i.idEquipo = %s
                UNION ALL
                SELECT traslado.fechaTraslado, traslado.idTraslado, "Traslado",
                    "Nombre", "observacion"
                FROM traslado, traslacion
                WHERE traslacion.idTraslado = traslado.idTraslado AND traslacion.idEquipo = %s
                UNION ALL
                SELECT a.fecha_inicioAsignacion, a.idAsignacion, "Asignacion",
                    a.ObservacionAsignacion, f.nombreFuncionario 
                FROM asignacion a
                INNER JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
                INNER JOIN equipo_asignacion ea on a.idAsignacion = ea.idAsignacion
                WHERE ea.idEquipo = %s
                UNION ALL
                SELECT a.fechaDevolucion, a.idAsignacion, "Devolucion",
                    a.ObservacionAsignacion, f.nombreFuncionario
                FROM asignacion a
                INNER JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
                INNER JOIN equipo_asignacion ea on a.idAsignacion = ea.idAsignacion
                WHERE ea.idEquipo = %s
                ORDER BY fecha DESC
                """, (idEquipo, idEquipo, idEquipo, idEquipo))
    data_eventos = cur.fetchall()
    cur.execute(
        """
                SELECT e.idEquipo, e.Cod_inventarioEquipo, e.Num_serieEquipo, 
                e.ObservacionEquipo, e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, 
                e.numerotelefonicoEquipo,e.idEstado_Equipo, e.idUnidad, 
                e.idOrden_compra, e.idModelo_equipo,te.idTipo_equipo, te.nombreTipo_Equipo, 
                ee.idEstado_equipo, ee.nombreEstado_equipo, u.idUnidad, u.nombreUnidad, 
                oc.idOrden_compra, oc.nombreOrden_compra,
    moe.idModelo_equipo, moe.nombreModeloequipo
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    WHERE e.idEquipo = {}
                """.format(idEquipo))
    data_equipo = cur.fetchone()
    return render_template("equipo_detalles.html", equipo=data_equipo, eventos=data_eventos)

#
    #cur.execute("""
            #SELECT *
            #FROM (
                #SELECT *
                #FROM asignacion
                #WHERE asignacion.idEquipo = %s 
            #)
            #LEFT OUTER JOIN (
                #SELECT *
                #FROM devolucion
                #WHERE devolucion.idEquipo = %s
            #)
            #LEFT OUTER JOIN (
                #SELECT *
                #FROM incidencia
                #WHERE incidencia.idEquipo = %s
            #)
            #LEFT OUTER JOIN (

            #)
                #""")
@equipo.route("/test_excel_form", methods=["POST"])
@loguear_requerido
def test_excel_form():
    #para el uso de la pagina de otros
    tipos = ("aio", "notebook", "impresoras", "bam", "proyectores", "telefonos", "disco_duro",
             "tablets")
    todo_check = request.form.get('todo_check')
    #si se imprime todo en una hoja usar la funcion ya creada
    if(todo_check == "on"):
        print("test")
        return crear_excel()
    #de lo contrario imprimir cada hoja individualmente
    computadora_check = request.form.get('AIO_check')
    notebooks_check = request.form.get('Notebooks')
    impresoras_check = request.form.get('impresoras_check')
    bam_check = request.form.get('bam_check')
    proyectores_check = request.form.get('proyectores_check')
    telefonos_check = request.form.get('telefonos_check')
    HDD_check = request.form.get('HDD_check')
    tablets_check = request.form.get('tablets_check')
    otros_check = request.form.get('otros_check')
    wb = Workbook()
    ws = wb.active
    if computadora_check == "on":
        ws.title = "AIO"
        añadir_hoja_de_tipo("AIO", ws)
        ws = wb.create_sheet("sheet")
    if notebooks_check == "on":
        ws.title = "Notebooks"
        añadir_hoja_de_tipo("Notebooks", ws)
        ws = wb.create_sheet("sheet")
    if impresoras_check == "on":
        ws.title = "Impresoras"
        añadir_hoja_de_tipo("impresoras", ws)
        ws = wb.create_sheet("sheet")
    if bam_check == "on":
        ws.title = "Bam"
        añadir_hoja_de_tipo("bam", ws)
        ws = wb.create_sheet("sheet")
    if proyectores_check == "on":
        ws.title = "Proyectoes"
        añadir_hoja_de_tipo("proyectores", ws)
        ws = wb.create_sheet("sheet")
    if telefonos_check == "on":
        ws.title = "Telefonos"
        añadir_hoja_de_tipo("telefonos", ws)
        ws = wb.create_sheet("sheet")
    if HDD_check == "on":
        ws.title = "Disco Duro"
        añadir_hoja_de_tipo("disco_duro", ws)
        ws = wb.create_sheet("sheet")
    if tablets_check == "on":
        ws.title = "Tablets"
        añadir_hoja_de_tipo("tablets", ws)
        ws = wb.create_sheet("sheet")
    #al ser otro requiere una consulta distinta ya que tendria que ser distinto a
    #todas las categorias anteriores
    if otros_check == "on":
        print("otros")
        ws.title = "Otros"
        añadir_hoja_de_otros(tipos, ws)
        ws = wb.create_sheet("sheet")
    
    wb.save("test.xlsx")
    return redirect(url_for("equipo.Equipo"))

def añadir_hoja_de_otros(tipos, ws):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    cur = mysql.connection.cursor()
    query = """

    SELECT *
    FROM
    (
    SELECT e.idEquipo, e.Cod_inventarioEquipo, 
           e.Num_serieEquipo, e.ObservacionEquipo,
           e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, 
           e.numerotelefonicoEquipo,
           te.idTipo_equipo, 
           te.nombreTipo_Equipo as tipo_equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, 
           u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
           com.nombreComuna, pro.nombreProvincia,
    moe.idModelo_equipo, moe.nombreModeloequipo, "" as nombreFuncionario,
                me.nombreMarcaEquipo, mo.nombreModalidad,
            pr.nombreProveedor
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = moe.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    LEFT JOIN marca_equipo me on me.idMarca_Equipo = te.idMarca_Equipo
    LEFT JOIN modalidad mo on mo.idModalidad = u.idModalidad

    LEFT JOIN comuna com ON com.idComuna = u.idComuna
    LEFT JOIN provincia pro ON pro.idProvincia = com.idProvincia
    INNER JOIN proveedor pr ON pr.idProveedor = oc.idProveedor

    WHERE ee.nombreEstado_equipo NOT LIKE "EN USO"
    UNION 
    SELECT  e.idEquipo, e.Cod_inventarioEquipo, 
            e.Num_serieEquipo, e.ObservacionEquipo, 
            e.codigoproveedor_equipo, e.macEquipo, 
            e.imeiEquipo, e.numerotelefonicoEquipo,
            te.idTipo_equipo, te.nombreTipo_Equipo,
            ee.idEstado_equipo, ee.nombreEstado_equipo, u.idUnidad,
            u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
            moe.idModelo_equipo, moe.nombreModeloequipo, f.nombreFuncionario,
            com.nombreComuna, pro.nombreProvincia,
            me.nombreMarcaEquipo, mo.nombreModalidad,
            pr.nombreProveedor
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    LEFT JOIN marca_equipo me on me.idMarca_Equipo = te.idMarca_Equipo
    LEFT JOIN modalidad mo on mo.idModalidad = u.idModalidad

    INNER JOIN equipo_asignacion ea on ea.idEquipo = e.idEquipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    LEFT JOIN asignacion a on a.idAsignacion = ea.idAsignacion
    LEFT JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
    LEFT JOIN comuna com ON com.idComuna = u.idComuna
    LEFT JOIN provincia pro ON pro.idProvincia = com.idProvincia
    INNER JOIN proveedor pr ON pr.idProveedor = oc.idProveedor
    WHERE ee.nombreEstado_equipo LIKE "EN USO"
    ) as subquery
    WHERE

"""
    
    for i in range(0, len(tipos)):
        query += """
        tipo_equipo NOT LIKE '{}'
        """.format(tipos[i]) 
        if(i != len(tipos) - 1):
            query += " AND "
    cur.execute(query)
    equipo_data = cur.fetchall()

    encabezado = (["Provincia", "Comuna", "Modalidad", "Codigo Proveedor", "Nombre", "Tipo de Bien", "Marca", "Modelo", 
               "N° Serie", "Codigo Inventario", "Nombre Proveedor"])
    print("encabezado len: " +  str(len(encabezado)))
    print(encabezado[10])
    for i in range(0, len(encabezado)):
        print(i)
        char = chr(65 + i)
        ws[char + str(1)].fill = PatternFill(start_color="000ff000", fill_type = "solid")
        ws.column_dimensions[char].width = 20
        ws[char + str(1)] = encabezado[i]

    i = 0
    def fillCell(data, fila):
        nonlocal i
        char = chr(65 + i)
        i += 1
        ws[char + str(fila)] = data 
    for fila in range(0, len(equipo_data)):
        i = 0
        #65 = A en ASCII
        #consegir lista de valores y extraer la lista de valires en cada for interior
        fillCell(equipo_data[fila]['nombreProvincia'], fila + 2)
        fillCell(equipo_data[fila]['nombreComuna'], fila + 2)
        fillCell(equipo_data[fila]['nombreModalidad'], fila + 2)
        fillCell(equipo_data[fila]['codigoproveedor_equipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreUnidad'], fila + 2)
        fillCell(equipo_data[fila]['tipo_equipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreMarcaEquipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreModeloequipo'], fila + 2)
        fillCell(equipo_data[fila]['Num_serieEquipo'], fila + 2)
        fillCell(equipo_data[fila]['Cod_inventarioEquipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreProveedor'], fila + 2)



    pass


def añadir_hoja_de_tipo(tipo, ws):
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    cur = mysql.connection.cursor()
    cur.execute(""" 
    SELECT *
    FROM
    (
    SELECT e.idEquipo, e.Cod_inventarioEquipo, 
           e.Num_serieEquipo, e.ObservacionEquipo,
           e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, 
           e.numerotelefonicoEquipo,
           te.idTipo_equipo, 
           te.nombreTipo_Equipo as tipo_equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, 
           u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
           com.nombreComuna, pro.nombreProvincia,
    moe.idModelo_equipo, moe.nombreModeloequipo, "" as nombreFuncionario,
                me.nombreMarcaEquipo, mo.nombreModalidad,
                pr.nombreProveedor
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    LEFT JOIN marca_equipo me on me.idMarca_Equipo = te.idMarca_Equipo
    LEFT JOIN modalidad mo on mo.idModalidad = u.idModalidad

    LEFT JOIN comuna com ON com.idComuna = u.idComuna
    LEFT JOIN provincia pro ON pro.idProvincia = com.idProvincia
    INNER JOIN proveedor pr ON oc.idProveedor = pr.idProveedor

    WHERE ee.nombreEstado_equipo NOT LIKE "EN USO"
    UNION 
    SELECT  e.idEquipo, e.Cod_inventarioEquipo, 
            e.Num_serieEquipo, e.ObservacionEquipo, 
            e.codigoproveedor_equipo, e.macEquipo, 
            e.imeiEquipo, e.numerotelefonicoEquipo,
            te.idTipo_equipo, te.nombreTipo_Equipo,
            ee.idEstado_equipo, ee.nombreEstado_equipo, u.idUnidad,
            u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
            moe.idModelo_equipo, moe.nombreModeloequipo, f.nombreFuncionario,
            com.nombreComuna, pro.nombreProvincia,
            me.nombreMarcaEquipo, mo.nombreModalidad,
            pr.nombreProveedor
                
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    LEFT JOIN marca_equipo me on me.idMarca_Equipo = te.idMarca_Equipo
    LEFT JOIN modalidad mo on mo.idModalidad = u.idModalidad

    INNER JOIN equipo_asignacion ea on ea.idEquipo = e.idEquipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    LEFT JOIN asignacion a on a.idAsignacion = ea.idAsignacion
    LEFT JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
    LEFT JOIN comuna com ON com.idComuna = u.idComuna
    LEFT JOIN provincia pro ON pro.idProvincia = com.idProvincia
    INNER JOIN proveedor pr ON oc.idProveedor = pr.idProveedor
    WHERE ee.nombreEstado_equipo LIKE "EN USO"
    ) as subquery
    WHERE tipo_equipo LIKE %s
                """, (tipo,))
    equipo_data = cur.fetchall()

    encabezado = (["Provincia", "Comuna", "Modalidad", "Codigo Proveedor", "Nombre", 
                   "CodigoUnidad","Tipo de Bien", "Marca", "Modelo", 
               "N° Serie", "Codigo Inventario", "Nombre Proveedor"])
    for i in range(0, len(encabezado)):
        char = chr(65 + i)
        ws[char + str(1)].fill = PatternFill(start_color="000ff000", fill_type = "solid")
        ws.column_dimensions[char].width = 20
        ws[char + str(1)] = encabezado[i]

    i = 0
    def fillCell(data, fila):
        nonlocal i
        char = chr(65 + i)
        i += 1
        ws[char + str(fila)] = data 
    for fila in range(0, len(equipo_data)):
        i = 0
        #65 = A en ASCII
        #consegir lista de valores y extraer la lista de valires en cada for interior
        fillCell(equipo_data[fila]['nombreProvincia'], fila + 2)
        fillCell(equipo_data[fila]['nombreComuna'], fila + 2)
        fillCell(equipo_data[fila]['nombreModalidad'], fila + 2)
        fillCell(equipo_data[fila]['codigoproveedor_equipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreUnidad'], fila + 2)
        fillCell(equipo_data[fila]['idUnidad'], fila + 2)
        fillCell(equipo_data[fila]['tipo_equipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreMarcaEquipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreModeloequipo'], fila + 2)
        fillCell(equipo_data[fila]['Num_serieEquipo'], fila + 2)
        fillCell(equipo_data[fila]['Cod_inventarioEquipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreProveedor'], fila + 2)



    #ingresar datos
    return
#exportar a pdf
@equipo.route("/equipo/crear_excel")
@loguear_requerido
def crear_excel():
    if "user" not in session:
        flash("you are NOT authorized")
        return redirect("/ingresar")
    #buscar columnas
    wb = Workbook()
    ws = wb.active

    #consulta datos
    cur = mysql.connection.cursor()
    cur.execute(""" 
    SELECT *
    FROM
    (
    SELECT e.idEquipo, e.Cod_inventarioEquipo, 
           e.Num_serieEquipo, e.ObservacionEquipo,
           e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, 
           e.numerotelefonicoEquipo,
           te.idTipo_equipo, 
           te.nombreTipo_Equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, 
           u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
           com.nombreComuna, pro.nombreProvincia,
        moe.idModelo_equipo, moe.nombreModeloequipo, "" as nombreFuncionario,
                me.nombreMarcaEquipo, mo.nombreModalidad,
                pr.nombreProveedor
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    INNER JOIN proveedor pr ON oc.idProveedor = pr.idProveedor
    LEFT JOIN marca_equipo me on me.idMarca_Equipo = te.idMarca_Equipo
    LEFT JOIN modalidad mo on mo.idModalidad = u.idModalidad

    LEFT JOIN comuna com ON com.idComuna = u.idComuna
    LEFT JOIN provincia pro ON pro.idProvincia = com.idProvincia

    WHERE ee.nombreEstado_equipo NOT LIKE "EN USO"
    UNION 
    SELECT  e.idEquipo, e.Cod_inventarioEquipo, 
            e.Num_serieEquipo, e.ObservacionEquipo, 
            e.codigoproveedor_equipo, e.macEquipo, 
            e.imeiEquipo, e.numerotelefonicoEquipo,
            te.idTipo_equipo, te.nombreTipo_Equipo,
            ee.idEstado_equipo, ee.nombreEstado_equipo, u.idUnidad,
            u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
            moe.idModelo_equipo, moe.nombreModeloequipo, f.nombreFuncionario,
            com.nombreComuna, pro.nombreProvincia,
            me.nombreMarcaEquipo, mo.nombreModalidad,
            pr.nombreProveedor
    FROM equipo e
    INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
    INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
    INNER JOIN unidad u on u.idUnidad = e.idUnidad
    INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra
    LEFT JOIN marca_equipo me on me.idMarca_Equipo = te.idMarca_Equipo
    LEFT JOIN modalidad mo on mo.idModalidad = u.idModalidad

    INNER JOIN equipo_asignacion ea on ea.idEquipo = e.idEquipo
    INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
    LEFT JOIN asignacion a on a.idAsignacion = ea.idAsignacion
    LEFT JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
    LEFT JOIN comuna com ON com.idComuna = u.idComuna
    LEFT JOIN provincia pro ON pro.idProvincia = com.idProvincia
    INNER JOIN proveedor pr ON oc.idProveedor = pr.idProveedor
    WHERE ee.nombreEstado_equipo LIKE "EN USO"
    ) as subquery
    
                """)
    equipo_data = cur.fetchall()

    #generar encabezado
    #encabezado

    encabezado = (["Provincia", "Comuna", "Modalidad", "Codigo Proveedor", "Nombre", "Tipo de Bien", "Marca", "Modelo", 
               "N° Serie", "Codigo Inventario", "Nombre Proveedor"])
    for i in range(0, len(encabezado)):
        char = chr(65 + i)
        ws[char + str(1)].fill = PatternFill(start_color="000ff000", fill_type = "solid")
        ws.column_dimensions[char].width = 20
        ws[char + str(1)] = encabezado[i]

    i = 0
    def fillCell(data, fila):
        nonlocal i
        char = chr(65 + i)
        i += 1
        ws[char + str(fila)] = data 
    for fila in range(0, len(equipo_data)):
        i = 0
        #65 = A en ASCII
        #consegir lista de valores y extraer la lista de valires en cada for interior
        fillCell(equipo_data[fila]['nombreProvincia'], fila + 2)
        fillCell(equipo_data[fila]['nombreComuna'], fila + 2)
        fillCell(equipo_data[fila]['nombreModalidad'], fila + 2)
        fillCell(equipo_data[fila]['codigoproveedor_equipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreUnidad'], fila + 2)
        fillCell(equipo_data[fila]['nombreTipo_Equipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreMarcaEquipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreModeloequipo'], fila + 2)
        fillCell(equipo_data[fila]['Num_serieEquipo'], fila + 2)
        fillCell(equipo_data[fila]['Cod_inventarioEquipo'], fila + 2)
        fillCell(equipo_data[fila]['nombreProveedor'], fila + 2)




    #ingresar datos
    wb.save("test.xlsx")
    return redirect(url_for("equipo.Equipo"))

def crear_pagina_todojunto(wb):
    return wb

@equipo.route("/equipo/importar_excel")
@administrador_requerido
def importar_excel(url):
    #Nesesito el excel para ver el formato
    pass

#buscar un equipo singular por id
@equipo.route("/equipo/buscar_equipo/<id>")
@loguear_requerido
def buscar_equipo(id):
    cur = mysql.connection.cursor()
    cur.execute("""
            SELECT *
                FROM
                (
                SELECT e.idEquipo, e.Cod_inventarioEquipo, 
                    e.Num_serieEquipo, e.ObservacionEquipo,
                    e.codigoproveedor_equipo, e.macEquipo, e.imeiEquipo, 
                    e.numerotelefonicoEquipo,
                    te.idTipo_equipo, 
                    te.nombreTipo_Equipo, ee.idEstado_equipo, ee.nombreEstado_equipo, 
                    u.idUnidad, u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
                moe.idModelo_equipo, moe.nombreModeloequipo, "" as nombreFuncionario
                FROM equipo e
                INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
                INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
                INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
                INNER JOIN unidad u on u.idUnidad = e.idUnidad
                INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra

                WHERE ee.nombreEstado_equipo NOT LIKE "EN USO"
                UNION 
                SELECT  e.idEquipo, e.Cod_inventarioEquipo, 
                        e.Num_serieEquipo, e.ObservacionEquipo, 
                        e.codigoproveedor_equipo, e.macEquipo, 
                        e.imeiEquipo, e.numerotelefonicoEquipo,
                        te.idTipo_equipo, te.nombreTipo_Equipo,
                        ee.idEstado_equipo, ee.nombreEstado_equipo, u.idUnidad,
                        u.nombreUnidad, oc.idOrden_compra, oc.nombreOrden_compra,
                        moe.idModelo_equipo, moe.nombreModeloequipo, f.nombreFuncionario
                FROM equipo e
                INNER JOIN modelo_equipo moe on moe.idModelo_Equipo = e.idModelo_equipo
                INNER JOIN tipo_equipo te on te.idTipo_equipo = moe.idTipo_Equipo
                INNER JOIN unidad u on u.idUnidad = e.idUnidad
                INNER JOIN orden_compra oc on oc.idOrden_compra = e.idOrden_compra

                INNER JOIN equipo_asignacion ea on ea.idEquipo = e.idEquipo
                INNER JOIN estado_equipo ee on ee.idEstado_equipo = e.idEstado_Equipo
                INNER JOIN asignacion a on a.idAsignacion = ea.idAsignacion
                INNER JOIN funcionario f on f.rutFuncionario = a.rutFuncionario
                WHERE ee.nombreEstado_equipo LIKE "EN USO" 
                ) as subquery
                WHERE idEquipo = %s

    """, (id,))
    Equipos = cur.fetchall()
    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    return render_template(
        "equipo.html",
        equipo=Equipos,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=1,
        lastpage=True,
    )

#buscar todos los equipos en base a una palabra de busqueda
@equipo.route("/consulta_equipo", methods =["POST"])
@equipo.route("/consulta_equipo/<page>", methods =["POST"])
@loguear_requerido
def consulta_equipo(page = 1):
    palabra = request.form["palabra"]
    if palabra == "":
        print("error_redirect")
    page = int(page)
    perpage = getPerPage()
    offset = (int(page) - 1) * perpage
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM equipo")
    total = cur.fetchone()
    total = int(str(total).split(":")[1].split("}")[0])
    cur = mysql.connection.cursor()
    query = f"""
    set palabra = CONVERT('%{palabra}%' USING utf8)
    SELECT *
    FROM superequipo se
    WHERE se.Cod_inventarioEquipo LIKE palabra OR
    se.Num_serieEquipo LIKE '%{palabra}%' OR
    se.codigoproveedor_equipo LIKE '%{palabra}%' OR
    se.nombreidTipoequipo LIKE '%{palabra}%' OR
    se.nombreEstado_equipo LIKE '%{palabra}%' OR
    se.idUnidad LIKE '%{palabra}%' OR
    se.nombreUnidad LIKE '%{palabra}%' OR
    se.nombreOrden_compra LIKE '%{palabra}%' OR
    se.nombreModeloequipo LIKE '%{palabra}%' OR
    se.nombreFuncionario LIKE '%{palabra}%'
    LIMIT {perpage} OFFSET {offset}
    """
    print(query)
    cur.execute(query)
    equipos = cur.fetchall()

    cur.execute("SELECT * FROM tipo_equipo")
    tipoe_data = cur.fetchall()
    cur.execute("SELECT idEstado_equipo, nombreEstado_equipo FROM estado_equipo")
    estadoe_data = cur.fetchall()
    cur.execute("SELECT idUnidad, nombreUnidad FROM Unidad")
    ubi_data = cur.fetchall()
    cur.execute("SELECT idOrden_compra, nombreOrden_compra FROM orden_compra")
    ordenc_data = cur.fetchall()
    cur.execute("SELECT idModelo_Equipo, nombreModeloequipo FROM modelo_equipo")
    modeloe_data = cur.fetchall()

    return render_template(
        "equipo.html",
        equipo=equipos,
        tipo_equipo=tipoe_data,
        estado_equipo=estadoe_data,
        orden_compra=ordenc_data,
        Unidad=ubi_data,
        modelo_equipo=modeloe_data,
        page=page,
        lastpage=page < (total / perpage) + 1,
    )